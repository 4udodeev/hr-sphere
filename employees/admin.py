from django.contrib import admin
from django.shortcuts import redirect, render
from django.urls import path
from openpyxl.reader.excel import load_workbook
from hr_sphere.forms import XlsxImportForm
from django.contrib.auth.base_user import BaseUserManager
from datetime import date

from .models import *

admin.site.register(Organization)
# admin.site.register(Subdivision)
# admin.site.register(Position)
# admin.site.register(Employee)

class SubdivisionAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'parent_subdivision', 'org_id')
    change_list_template = 'hr_sphere/record_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        # Добавляем URL нашего обработчика импорта.
        my_urls = [
            path('import-records-from-xlsx/', self.import_records_from_xlsx),
        ]
        return my_urls + urls

    def import_records_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']

            workbook = load_workbook(filename=xlsx_file, read_only=True)
            worksheet = workbook.active

            # Читаем файл построчно и создаем объекты.
            i = 0
            for row in worksheet.rows:
                records_to_save = []
                if not Subdivision.objects.filter(code=row[0].value).exists():
                    new_obj = Subdivision(
                        code=row[0].value, 
                        name=row[1].value,  
                        org_id=Organization.objects.get(id=1),
                    )
                    records_to_save.append(new_obj)
                    Subdivision.objects.bulk_create(records_to_save)
                    i += 1
            
            for row in worksheet.rows:
                if Subdivision.objects.filter(code=row[0].value).exists():
                    cur_obj = Subdivision.objects.get(code=row[0].value)
                    if Subdivision.objects.filter(code=row[2].value).exists():
                        cur_obj.parent_subdivision = Subdivision.objects.get(code=row[2].value)
                    cur_obj.save()

            self.message_user(request, f'Импортировано строк: {i}.')
            return redirect('/admin/employees/subdivision/')

        context['form'] = XlsxImportForm()
        return render(request, 'hr_sphere/add_records_form.html', context=context)


admin.site.register(Subdivision, SubdivisionAdmin)


class PositionAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'is_boss', 'employee', 'subdivision', 'org', 'base_position')
    change_list_template = 'hr_sphere/record_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        # Добавляем URL нашего обработчика импорта.
        my_urls = [
            path('import-records-from-xlsx/', self.import_records_from_xlsx),
        ]
        return my_urls + urls

    def import_records_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']

            workbook = load_workbook(filename=xlsx_file, read_only=True)
            worksheet = workbook.active

            # Читаем файл построчно и создаем объекты.
            i = 0
            for row in worksheet.rows:
                records_to_save = []
                if not Position.objects.filter(code=row[0].value).exists():
                    new_obj = Position(
                        code=row[0].value, 
                        name=row[11].value,
                        is_boss=False,
                        employee=Employee.objects.get(code=row[0].value),
                        subdivision=Subdivision.objects.get(code=row[12].value) if Subdivision.objects.filter(code=row[10].value).exists() else None,
                        org=Organization.objects.get(id=1),
                    )
                    records_to_save.append(new_obj)
                    Position.objects.bulk_create(records_to_save)
                    i += 1
            
            j=0
            for row in worksheet.rows:
                if Position.objects.filter(code=row[0].value).exists():
                    cur_obj = Position.objects.get(code=row[0].value)
                    cur_obj.name=row[11].value,
                    cur_obj.is_boss=False,
                    cur_obj.employee=Employee.objects.get(code=row[0].value),
                    cur_obj.subdivision=Subdivision.objects.get(code=row[12].value),
                    cur_obj.org=Organization.objects.get(id=1),
                    cur_obj.save()
                    j += 1

            self.message_user(request, f'Импортировано строк: {i}.')
            self.message_user(request, f'Изменено строк: {j}.')
            return redirect('/admin/employees/position/')

        context['form'] = XlsxImportForm()
        return render(request, 'hr_sphere/add_records_form.html', context=context)


admin.site.register(Position, PositionAdmin)


class EmployeeAdmin(admin.ModelAdmin):
    list_display = ('code', 'fullname', 'sex', 'birth_date', 'email', 'login', 'hire_date', 'is_dismiss', 'dismiss_date')
    change_list_template = 'hr_sphere/record_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        # Добавляем URL нашего обработчика импорта.
        my_urls = [
            path('import-records-from-xlsx/', self.import_records_from_xlsx),
        ]
        return my_urls + urls

    def import_records_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']

            workbook = load_workbook(filename=xlsx_file, read_only=True)
            worksheet = workbook.active

            # Читаем файл построчно и создаем объекты.
            i = 0
            for row in worksheet.rows:
                records_to_save = []
                if not Employee.objects.filter(code=row[0].value).exists():
                    new_obj = Employee(
                        code=row[0].value,
                        lastname = row[1].value,
                        firstname = row[2].value,
                        middlename = row[3].value,
                        fullname = f'{row[1].value} {row[2].value} {row[3].value}'.rstrip(),
                        sex = row[4].value,
                        birth_date = row[5].value.date(),
                        phone = row[6].value,
                        email = row[7].value,
                        education = row[8].value,
                        login = row[0].value,
                        password = BaseUserManager().make_random_password(8),
                        change_password = True,
                        hire_date = row[16].value.date(),
                        is_dismiss = True if row[18].value=='+' else False,
                        is_banned = True if row[18].value=='+' else False,
                        dismiss_date = row[17].value.date() if row[18].value=='+' else None,
                    )
                    records_to_save.append(new_obj)
                    Employee.objects.bulk_create(records_to_save)
                    i += 1
            
            j=0
            for row in worksheet.rows:
                if Employee.objects.filter(code=row[0].value).exists():
                    cur_obj = Employee.objects.get(code=row[0].value)
                    cur_obj.lastname = row[1].value
                    cur_obj.firstname = row[2].value
                    cur_obj.middlename = row[3].value
                    cur_obj.fullname = f'{row[1].value} {row[2].value} {row[3].value}'.rstrip()
                    cur_obj.sex = row[4].value
                    cur_obj.birth_date = row[5].value.date()
                    cur_obj.phone = row[6].value
                    cur_obj.email = row[7].value
                    cur_obj.education = row[8].value
                    cur_obj.login = row[0].value
                    cur_obj.hire_date = row[16].value.date()
                    cur_obj.is_dismiss = True if row[18].value=='+' else False
                    cur_obj.is_banned = True if row[18].value=='+' else False
                    cur_obj.dismiss_date = row[17].value.date() if row[18].value=='+' else None
                    cur_obj.save()
                    j += 1

            self.message_user(request, f'Импортировано строк: {i}')
            self.message_user(request, f'Изменено строк: {j}.')
            return redirect('/admin/employees/employee/')

        context['form'] = XlsxImportForm()
        return render(request, 'hr_sphere/add_records_form.html', context=context)


admin.site.register(Employee, EmployeeAdmin)